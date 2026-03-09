"""Excel 帳票出力 テスト (Issue #21, #22)"""

from pathlib import Path

from openpyxl import load_workbook

from src.engine.judgment import JudgmentDetail, JudgmentResult, JudgmentType
from src.engine.test_runner import ExecutionSummary, TestResult, TestStatus
from src.report.excel_report import ExcelReportGenerator


def _make_summary() -> ExecutionSummary:
    return ExecutionSummary(
        total=3,
        passed=2,
        failed=1,
        start_time="2026-03-09T10:00:00",
        end_time="2026-03-09T10:05:00",
        config_file="test.cfg",
        results=[
            TestResult("TC-001", "テスト1", TestStatus.PASSED,
                       actual_value="100", expected_value="100",
                       start_time="2026-03-09T10:00:00", duration_ms=1000.0),
            TestResult("TC-002", "テスト2", TestStatus.PASSED,
                       actual_value="200", expected_value="200",
                       start_time="2026-03-09T10:01:00", duration_ms=1500.0),
            TestResult("TC-003", "テスト3", TestStatus.FAILED,
                       actual_value="999", expected_value="300",
                       start_time="2026-03-09T10:03:00", duration_ms=500.0),
        ],
    )


def _make_judgments() -> list[JudgmentDetail]:
    return [
        JudgmentDetail(
            test_case_id="TC-001", result=JudgmentResult.OK,
            judgment_type=JudgmentType.EXACT, signal_name="EngineSpeed",
            expected_value="100", actual_value="100",
        ),
        JudgmentDetail(
            test_case_id="TC-002", result=JudgmentResult.OK,
            judgment_type=JudgmentType.RANGE, signal_name="VehicleSpeed",
            expected_value="0 ~ 300", actual_value="200",
        ),
        JudgmentDetail(
            test_case_id="TC-003", result=JudgmentResult.NG,
            judgment_type=JudgmentType.EXACT, signal_name="BrakeStatus",
            expected_value="300", actual_value="999", difference="699",
            reason="差異 699 が許容差 0 を超えています",
        ),
    ]


class TestExcelReportSummarySheet:
    def test_summary_sheet_created(self, tmp_path: Path) -> None:
        gen = ExcelReportGenerator()
        output = tmp_path / "report.xlsx"
        gen.generate(_make_summary(), output_path=output)
        wb = load_workbook(output)
        assert "サマリ" in wb.sheetnames

    def test_summary_sheet_has_header_info(self, tmp_path: Path) -> None:
        gen = ExcelReportGenerator()
        output = tmp_path / "report.xlsx"
        gen.generate(_make_summary(), output_path=output)
        wb = load_workbook(output)
        ws = wb["サマリ"]
        assert ws["A1"].value == "テスト結果サマリ"
        assert ws["A3"].value == "実行日時:"

    def test_summary_sheet_has_results(self, tmp_path: Path) -> None:
        gen = ExcelReportGenerator()
        output = tmp_path / "report.xlsx"
        gen.generate(_make_summary(), output_path=output)
        wb = load_workbook(output)
        ws = wb["サマリ"]
        # ヘッダー行 (row 7) + データ3行
        assert ws.cell(row=8, column=1).value == "TC-001"
        assert ws.cell(row=10, column=1).value == "TC-003"

    def test_summary_sheet_result_coloring(self, tmp_path: Path) -> None:
        gen = ExcelReportGenerator()
        output = tmp_path / "report.xlsx"
        gen.generate(_make_summary(), output_path=output)
        wb = load_workbook(output)
        ws = wb["サマリ"]
        # PASSED のセルは緑、FAILED のセルは赤
        passed_cell = ws.cell(row=8, column=6)
        failed_cell = ws.cell(row=10, column=6)
        assert passed_cell.fill.start_color.rgb == "00C6EFCE"
        assert failed_cell.fill.start_color.rgb == "00FFC7CE"


class TestExcelReportStatisticsSheet:
    def test_statistics_sheet_created(self, tmp_path: Path) -> None:
        gen = ExcelReportGenerator()
        output = tmp_path / "report.xlsx"
        gen.generate(_make_summary(), output_path=output)
        wb = load_workbook(output)
        assert "統計" in wb.sheetnames

    def test_statistics_has_counts(self, tmp_path: Path) -> None:
        gen = ExcelReportGenerator()
        output = tmp_path / "report.xlsx"
        gen.generate(_make_summary(), output_path=output)
        wb = load_workbook(output)
        ws = wb["統計"]
        # テスト総数
        assert ws.cell(row=3, column=1).value == "テスト総数"
        assert ws.cell(row=3, column=2).value == 3
        # OK数
        assert ws.cell(row=4, column=1).value == "OK数"
        assert ws.cell(row=4, column=2).value == 2

    def test_statistics_pass_rate(self, tmp_path: Path) -> None:
        gen = ExcelReportGenerator()
        output = tmp_path / "report.xlsx"
        gen.generate(_make_summary(), output_path=output)
        wb = load_workbook(output)
        ws = wb["統計"]
        # 合格率行
        assert ws.cell(row=9, column=1).value == "合格率"
        assert ws.cell(row=9, column=2).value == "66.7%"


class TestExcelReportDetailSheet:
    def test_detail_sheets_created(self, tmp_path: Path) -> None:
        gen = ExcelReportGenerator()
        output = tmp_path / "report.xlsx"
        gen.generate(_make_summary(), judgments=_make_judgments(), output_path=output)
        wb = load_workbook(output)
        assert "TC-001" in wb.sheetnames
        assert "TC-002" in wb.sheetnames
        assert "TC-003" in wb.sheetnames

    def test_detail_sheet_content(self, tmp_path: Path) -> None:
        gen = ExcelReportGenerator()
        output = tmp_path / "report.xlsx"
        gen.generate(_make_summary(), judgments=_make_judgments(), output_path=output)
        wb = load_workbook(output)
        ws = wb["TC-001"]
        assert "TC-001" in str(ws["A1"].value)
        # 判定結果
        assert ws.cell(row=9, column=1).value == "判定結果"
        assert ws.cell(row=9, column=2).value == "OK"

    def test_detail_sheet_ng_coloring(self, tmp_path: Path) -> None:
        gen = ExcelReportGenerator()
        output = tmp_path / "report.xlsx"
        gen.generate(_make_summary(), judgments=_make_judgments(), output_path=output)
        wb = load_workbook(output)
        ws = wb["TC-003"]
        # 判定結果セル (row 9)
        result_cell = ws.cell(row=9, column=2)
        assert result_cell.value == "NG"
        assert result_cell.fill.start_color.rgb == "00FFC7CE"

    def test_no_detail_sheets_without_judgments(self, tmp_path: Path) -> None:
        gen = ExcelReportGenerator()
        output = tmp_path / "report.xlsx"
        gen.generate(_make_summary(), output_path=output)
        wb = load_workbook(output)
        assert "TC-001" not in wb.sheetnames

    def test_generate_default_path(self, tmp_path: Path) -> None:
        """出力パス未指定時にデフォルトパスで生成"""
        import os
        os.chdir(tmp_path)
        gen = ExcelReportGenerator()
        output = gen.generate(_make_summary())
        assert output.exists()
        assert output.suffix == ".xlsx"
