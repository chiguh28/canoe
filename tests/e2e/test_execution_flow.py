"""実行→判定→帳票 E2E テスト

Phase 3→4 のワークフロー:
確定済みパラメータ → TestRunner → JudgmentEngine → ExcelReportGenerator
"""

from __future__ import annotations

import json
from pathlib import Path

import pytest
from openpyxl import load_workbook

from src.engine.judgment import (
    ChangeDirection,
    JudgmentCriteria,
    JudgmentDetail,
    JudgmentEngine,
    JudgmentResult,
    JudgmentType,
)
from src.engine.log_manager import LogEntry, LogManager, TestLog
from src.engine.test_runner import TestRunner
from src.models.test_pattern import TestPattern
from src.report.excel_report import ExcelReportGenerator

from .conftest import MockCANoeCOM


@pytest.mark.e2e
class TestExecutionFlow:
    """E2E-N04〜N07: テスト実行→判定→帳票"""

    def test_execute_judge_report(
        self,
        mock_canoe: MockCANoeCOM,
        tmp_path: Path,
    ) -> None:
        """E2E-N04: 確定済み JSON → 実行 → 判定 → Excel 帳票"""
        patterns = [
            TestPattern(
                test_case_id="TC-001",
                test_case_name="エンジン回転数テスト",
                target_signal="EngineSpeed",
                operation="回転数を 2000rpm に設定",
                expected_value="スロットル 20% 以上",
                wait_time_ms=0,
            ),
            TestPattern(
                test_case_id="TC-002",
                test_case_name="ブレーキテスト",
                target_signal="BrakePedal",
                operation="ブレーキ 50% 踏込",
                expected_value="ランプ点灯",
                wait_time_ms=0,
            ),
        ]

        # --- 実行 ---
        runner = TestRunner(com_wrapper=mock_canoe)
        summary = runner.execute(patterns, config_file="test.cfg")
        assert summary.total == 2

        # --- ログ作成（判定用） ---
        log_mgr = LogManager(log_dir=tmp_path / "logs")

        log_mgr.start_log("TC-001")
        log_mgr.add_entry(
            "TC-001",
            LogEntry(0.0, 1, "EngineData", "ThrottlePosition", 22.5),
        )
        log1 = log_mgr.end_log("TC-001")

        log_mgr.start_log("TC-002")
        log_mgr.add_entry(
            "TC-002",
            LogEntry(0.0, 1, "BrakeData", "BrakeLamp", 1.0),
        )
        log2 = log_mgr.end_log("TC-002")

        # --- 判定 ---
        engine = JudgmentEngine()
        j1 = engine.judge(
            "TC-001",
            log1,
            JudgmentCriteria(
                judgment_type=JudgmentType.RANGE,
                signal_name="ThrottlePosition",
                range_min=20.0,
                range_max=100.0,
            ),
        )
        j2 = engine.judge(
            "TC-002",
            log2,
            JudgmentCriteria(
                judgment_type=JudgmentType.EXACT,
                signal_name="BrakeLamp",
                expected_value=1.0,
                tolerance=0.0,
            ),
        )
        assert j1.result == JudgmentResult.OK
        assert j2.result == JudgmentResult.OK

        # --- 帳票出力 ---
        report_path = tmp_path / "report.xlsx"
        gen = ExcelReportGenerator()
        result_path = gen.generate(
            summary=summary,
            judgments=[j1, j2],
            output_path=report_path,
            config_file="test.cfg",
        )
        assert result_path.exists()

    def test_all_judgment_types(
        self,
        tmp_path: Path,
    ) -> None:
        """E2E-N05: 全判定タイプ混合テスト（EXACT/RANGE/CHANGE/TIMEOUT/COMPOUND）"""
        engine = JudgmentEngine()

        # --- EXACT 判定 ---
        exact_log = TestLog(
            test_case_id="TC-EXACT",
            entries=[LogEntry(0.0, 1, "Msg", "Sig", 100.0)],
        )
        exact_result = engine.judge(
            "TC-EXACT",
            exact_log,
            JudgmentCriteria(
                judgment_type=JudgmentType.EXACT,
                signal_name="Sig",
                expected_value=100.0,
                tolerance=0.5,
            ),
        )
        assert exact_result.result == JudgmentResult.OK

        # --- RANGE 判定 ---
        range_log = TestLog(
            test_case_id="TC-RANGE",
            entries=[LogEntry(0.0, 1, "Msg", "Temp", 85.0)],
        )
        range_result = engine.judge(
            "TC-RANGE",
            range_log,
            JudgmentCriteria(
                judgment_type=JudgmentType.RANGE,
                signal_name="Temp",
                range_min=80.0,
                range_max=100.0,
            ),
        )
        assert range_result.result == JudgmentResult.OK

        # --- CHANGE 判定 ---
        change_log = TestLog(
            test_case_id="TC-CHANGE",
            entries=[
                LogEntry(0.0, 1, "Msg", "Speed", 0.0),
                LogEntry(1.0, 1, "Msg", "Speed", 50.0),
            ],
        )
        change_result = engine.judge(
            "TC-CHANGE",
            change_log,
            JudgmentCriteria(
                judgment_type=JudgmentType.CHANGE,
                signal_name="Speed",
                change_direction=ChangeDirection.INCREASE,
            ),
        )
        assert change_result.result == JudgmentResult.OK

        # --- TIMEOUT 判定 ---
        timeout_log = TestLog(
            test_case_id="TC-TIMEOUT",
            entries=[
                LogEntry(0.0, 1, "Msg", "Response", 0.0),
                LogEntry(0.5, 1, "Msg", "Response", 1.0),
            ],
        )
        timeout_result = engine.judge(
            "TC-TIMEOUT",
            timeout_log,
            JudgmentCriteria(
                judgment_type=JudgmentType.TIMEOUT,
                signal_name="Response",
                timeout_ms=1000.0,
            ),
        )
        assert timeout_result.result == JudgmentResult.OK

        # --- COMPOUND 判定 (AND) ---
        compound_log = TestLog(
            test_case_id="TC-COMPOUND",
            entries=[
                LogEntry(0.0, 1, "Msg", "SigA", 100.0),
                LogEntry(0.0, 1, "Msg", "SigB", 50.0),
            ],
        )
        compound_result = engine.judge(
            "TC-COMPOUND",
            compound_log,
            JudgmentCriteria(
                judgment_type=JudgmentType.COMPOUND,
                signal_name="Compound",
                sub_criteria=[
                    JudgmentCriteria(
                        judgment_type=JudgmentType.EXACT,
                        signal_name="SigA",
                        expected_value=100.0,
                    ),
                    JudgmentCriteria(
                        judgment_type=JudgmentType.RANGE,
                        signal_name="SigB",
                        range_min=40.0,
                        range_max=60.0,
                    ),
                ],
                compound_operator="AND",
            ),
        )
        assert compound_result.result == JudgmentResult.OK

    def test_result_persistence(
        self,
        mock_canoe: MockCANoeCOM,
        tmp_path: Path,
    ) -> None:
        """E2E-N06: 結果 JSON + ログ CSV/JSON の永続化検証"""
        patterns = [
            TestPattern(
                test_case_id="TC-PERSIST",
                test_case_name="永続化テスト",
                operation="信号設定",
                expected_value="期待値",
                wait_time_ms=0,
            ),
        ]

        runner = TestRunner(com_wrapper=mock_canoe)
        summary = runner.execute(patterns)

        # 実行結果 JSON 保存 → 再読み込み
        results_json = tmp_path / "results.json"
        runner.save_results_json(results_json, summary)
        assert results_json.exists()

        loaded = json.loads(results_json.read_text(encoding="utf-8"))
        assert loaded["total"] == 1
        assert len(loaded["results"]) == 1

        # ログ CSV 保存 → 再読み込み
        log_mgr = LogManager(log_dir=tmp_path / "logs")
        log_mgr.start_log("TC-PERSIST")
        log_mgr.add_entry(
            "TC-PERSIST",
            LogEntry(0.0, 1, "Msg", "Sig", 42.0),
        )
        log_mgr.end_log("TC-PERSIST")

        csv_path = log_mgr.save_log_csv("TC-PERSIST")
        assert csv_path.exists()

        reloaded_csv = log_mgr.load_log_csv(csv_path)
        assert len(reloaded_csv.entries) == 1
        assert reloaded_csv.entries[0].value == 42.0

        # ログ JSON 保存 → 再読み込み
        json_path = log_mgr.save_log_json("TC-PERSIST")
        assert json_path.exists()

        reloaded_json = log_mgr.load_log_json(json_path)
        assert len(reloaded_json.entries) == 1
        assert reloaded_json.entries[0].value == 42.0

        # 判定結果 JSON 保存
        engine = JudgmentEngine()
        detail = engine.judge(
            "TC-PERSIST",
            reloaded_json,
            JudgmentCriteria(
                judgment_type=JudgmentType.EXACT,
                signal_name="Sig",
                expected_value=42.0,
            ),
        )
        judgment_json = tmp_path / "judgments.json"
        engine.save_results_json([detail], judgment_json)
        assert judgment_json.exists()

        judgment_data = json.loads(judgment_json.read_text(encoding="utf-8"))
        assert len(judgment_data) == 1
        assert judgment_data[0]["result"] == "OK"

    def test_excel_report_format(
        self,
        mock_canoe: MockCANoeCOM,
        tmp_path: Path,
    ) -> None:
        """E2E-N07: Excel 帳票フォーマット検証（3シート構成）"""
        patterns = [
            TestPattern(
                test_case_id="TC-RPT-001",
                test_case_name="帳票テスト1",
                operation="操作1",
                expected_value="期待値1",
                wait_time_ms=0,
            ),
            TestPattern(
                test_case_id="TC-RPT-002",
                test_case_name="帳票テスト2",
                operation="操作2",
                expected_value="期待値2",
                wait_time_ms=0,
            ),
            TestPattern(
                test_case_id="TC-RPT-003",
                test_case_name="帳票テスト3",
                operation="操作3",
                expected_value="期待値3",
                wait_time_ms=0,
            ),
        ]

        runner = TestRunner(com_wrapper=mock_canoe)
        summary = runner.execute(patterns, config_file="report_test.cfg")

        # 判定結果を作成
        judgments: list[JudgmentDetail] = []
        for tc_id in ["TC-RPT-001", "TC-RPT-002", "TC-RPT-003"]:
            judgments.append(
                JudgmentDetail(
                    test_case_id=tc_id,
                    result=JudgmentResult.OK if tc_id != "TC-RPT-003" else JudgmentResult.NG,
                    judgment_type=JudgmentType.EXACT,
                    signal_name="TestSignal",
                    expected_value="100",
                    actual_value="100" if tc_id != "TC-RPT-003" else "50",
                )
            )

        report_path = tmp_path / "format_test.xlsx"
        gen = ExcelReportGenerator()
        gen.generate(
            summary=summary,
            judgments=judgments,
            output_path=report_path,
        )
        assert report_path.exists()

        # Excel ファイルを開いてフォーマット検証
        wb = load_workbook(str(report_path))

        # 3シート + 各 TC-ID の詳細シート = サマリ + 統計 + 3詳細 = 5シート
        assert "サマリ" in wb.sheetnames
        assert "統計" in wb.sheetnames
        assert len(wb.sheetnames) >= 3  # サマリ + 統計 + 詳細シート

        # サマリシート検証
        ws_summary = wb["サマリ"]
        assert ws_summary["A1"].value == "テスト結果サマリ"

        # 統計シート検証
        ws_stats = wb["統計"]
        assert ws_stats["A1"].value == "テスト統計情報"

        wb.close()
