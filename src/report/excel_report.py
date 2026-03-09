"""Excel 帳票出力 (Issue #21, #22)

テスト結果を Excel 形式で出力する。
サマリシート、統計シート、詳細シートを生成する。
"""

from __future__ import annotations

import logging
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from src.engine.judgment import JudgmentDetail, JudgmentResult
from src.engine.test_runner import ExecutionSummary, TestResult, TestStatus

logger = logging.getLogger(__name__)

# スタイル定義
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
OK_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
NG_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
ERROR_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
OK_FONT = Font(color="006100")
NG_FONT = Font(color="9C0006")
ERROR_FONT = Font(color="9C6500")
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
TITLE_FONT = Font(bold=True, size=14)
SUBTITLE_FONT = Font(bold=True, size=12)


class ExcelReportGenerator:
    """Excel 帳票生成クラス"""

    def __init__(self) -> None:
        self._wb: Workbook | None = None
        self._tool_version: str = "0.1.0"

    def generate(
        self,
        summary: ExecutionSummary,
        judgments: list[JudgmentDetail] | None = None,
        output_path: Path | None = None,
        config_file: str = "",
    ) -> Path:
        """Excel 帳票を生成

        Args:
            summary: 実行サマリ
            judgments: 判定結果リスト
            output_path: 出力先パス
            config_file: CANoe 構成ファイル名

        Returns:
            生成された Excel ファイルのパス
        """
        self._wb = Workbook()

        # サマリシート
        self._create_summary_sheet(summary, config_file)

        # 統計シート
        self._create_statistics_sheet(summary)

        # 詳細シート（各テストケース）
        if judgments:
            for judgment in judgments:
                result = self._find_result(summary, judgment.test_case_id)
                self._create_detail_sheet(judgment, result)

        # デフォルトシート削除（既に他のシートが作成済みの場合）
        if "Sheet" in self._wb.sheetnames and len(self._wb.sheetnames) > 1:
            del self._wb["Sheet"]

        # 保存
        if output_path is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_path = Path(f"test_report_{timestamp}.xlsx")

        self._wb.save(str(output_path))
        logger.info("Excel帳票を生成しました: %s", output_path)
        return output_path

    def _create_summary_sheet(
        self, summary: ExecutionSummary, config_file: str
    ) -> None:
        """サマリシートを作成"""
        ws = self._wb.create_sheet("サマリ", 0) if self._wb else None
        if ws is None:
            return

        # ヘッダー情報
        ws["A1"] = "テスト結果サマリ"
        ws["A1"].font = TITLE_FONT
        ws["A3"] = "実行日時:"
        ws["B3"] = summary.start_time
        ws["A4"] = "ツールバージョン:"
        ws["B4"] = self._tool_version
        ws["A5"] = "CANoe構成ファイル:"
        ws["B5"] = config_file or summary.config_file

        # テスト結果テーブル
        headers = [
            "TC-ID", "テストケース名", "対象信号", "期待値",
            "実測値", "判定結果", "実行日時",
        ]
        start_row = 7
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=start_row, column=col, value=header)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal="center")

        # データ行
        for i, result in enumerate(summary.results):
            row = start_row + 1 + i
            ws.cell(row=row, column=1, value=result.test_case_id).border = THIN_BORDER
            ws.cell(row=row, column=2, value=result.test_case_name).border = THIN_BORDER
            ws.cell(row=row, column=3, value="").border = THIN_BORDER  # 信号名は判定結果から
            ws.cell(row=row, column=4, value=result.expected_value).border = THIN_BORDER
            ws.cell(row=row, column=5, value=result.actual_value).border = THIN_BORDER

            # 判定結果セルのスタイル
            status_cell = ws.cell(row=row, column=6, value=result.status.value)
            status_cell.border = THIN_BORDER
            if result.status == TestStatus.PASSED:
                status_cell.fill = OK_FILL
                status_cell.font = OK_FONT
            elif result.status == TestStatus.FAILED:
                status_cell.fill = NG_FILL
                status_cell.font = NG_FONT
            elif result.status == TestStatus.ERROR:
                status_cell.fill = ERROR_FILL
                status_cell.font = ERROR_FONT

            ws.cell(row=row, column=7, value=result.start_time).border = THIN_BORDER

        # 列幅調整
        column_widths = [12, 30, 20, 20, 20, 12, 25]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width

    def _create_statistics_sheet(self, summary: ExecutionSummary) -> None:
        """統計シートを作成"""
        ws = self._wb.create_sheet("統計") if self._wb else None
        if ws is None:
            return

        ws["A1"] = "テスト統計情報"
        ws["A1"].font = TITLE_FONT

        # 統計データ
        stats = [
            ("テスト総数", summary.total),
            ("OK数", summary.passed),
            ("NG数", summary.failed),
            ("ERROR数", summary.error),
            ("SKIP数", summary.skipped),
            ("中断数", summary.aborted),
            ("合格率", f"{summary.pass_rate:.1f}%"),
            ("実行時間合計", f"{summary.total_duration_ms:.1f} ms"),
            ("開始日時", summary.start_time),
            ("終了日時", summary.end_time),
        ]

        start_row = 3
        for i, (label, value) in enumerate(stats):
            row = start_row + i
            label_cell = ws.cell(row=row, column=1, value=label)
            label_cell.font = Font(bold=True)
            label_cell.border = THIN_BORDER
            value_cell = ws.cell(row=row, column=2, value=value)
            value_cell.border = THIN_BORDER

            # OK/NG のカラーリング
            if label == "OK数" and summary.passed > 0:
                value_cell.fill = OK_FILL
                value_cell.font = OK_FONT
            elif label == "NG数" and summary.failed > 0:
                value_cell.fill = NG_FILL
                value_cell.font = NG_FONT

        # 列幅調整
        ws.column_dimensions["A"].width = 18
        ws.column_dimensions["B"].width = 25

    def _create_detail_sheet(
        self, judgment: JudgmentDetail, result: TestResult | None
    ) -> None:
        """詳細シートを作成（テストケースごと）"""
        if self._wb is None:
            return

        # シート名（最大31文字制限）
        sheet_name = judgment.test_case_id[:31]
        ws = self._wb.create_sheet(sheet_name)

        ws["A1"] = f"テストケース詳細: {judgment.test_case_id}"
        ws["A1"].font = TITLE_FONT

        # テストケース情報
        info = [
            ("テストケースID", judgment.test_case_id),
            ("判定種別", judgment.judgment_type.value),
            ("対象信号", judgment.signal_name),
            ("期待値", judgment.expected_value),
            ("実測値", judgment.actual_value),
            ("差異", judgment.difference),
            ("判定結果", judgment.result.value),
            ("判定理由", judgment.reason),
        ]

        if result:
            info.extend([
                ("開始時刻", result.start_time),
                ("終了時刻", result.end_time),
                ("実行時間", f"{result.duration_ms:.1f} ms"),
                ("ログファイル", result.log_file),
            ])

        start_row = 3
        for i, (label, value) in enumerate(info):
            row = start_row + i
            label_cell = ws.cell(row=row, column=1, value=label)
            label_cell.font = Font(bold=True)
            label_cell.border = THIN_BORDER
            value_cell = ws.cell(row=row, column=2, value=value)
            value_cell.border = THIN_BORDER

            # 判定結果のカラーリング
            if label == "判定結果":
                if judgment.result == JudgmentResult.OK:
                    value_cell.fill = OK_FILL
                    value_cell.font = OK_FONT
                elif judgment.result == JudgmentResult.NG:
                    value_cell.fill = NG_FILL
                    value_cell.font = NG_FONT
                elif judgment.result == JudgmentResult.ERROR:
                    value_cell.fill = ERROR_FILL
                    value_cell.font = ERROR_FONT

        # 列幅調整
        ws.column_dimensions["A"].width = 18
        ws.column_dimensions["B"].width = 50

    def _find_result(
        self, summary: ExecutionSummary, test_case_id: str
    ) -> TestResult | None:
        """サマリからテスト結果を検索"""
        for r in summary.results:
            if r.test_case_id == test_case_id:
                return r
        return None
